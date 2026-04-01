from __future__ import annotations

import base64
import io
import json
import logging
import os
import re
import uuid
import xml.etree.ElementTree as ET
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from PIL import Image, ImageEnhance, ImageOps
import pytesseract
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from database import close_db, get_db, init_db
from models import PatientTable, VisitTable

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")
UPLOADS_VISITS = ROOT_DIR / "uploads" / "visits"

# Create the main app with lifespan
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    UPLOADS_VISITS.mkdir(parents=True, exist_ok=True)
    yield
    await close_db()


app = FastAPI(title="Clinic Digitization API", lifespan=lifespan)
api_router = APIRouter(prefix="/api")


# ============== Pydantic models ==============


class PatientBase(BaseModel):
    name: str
    age: int
    gender: str
    phone: str
    location: Optional[str] = ""
    doctor_name: str = ""
    notes: str = ""


class PatientCreate(PatientBase):
    pass


class Patient(PatientBase):
    id: str
    created_at: datetime

    class Config:
        from_attributes = False


class VisitCreate(BaseModel):
    patient_id: str
    symptoms: str = ""
    diagnosis: str = ""
    prescription: str = ""
    raw_text: str = ""
    structured_data: dict = Field(default_factory=dict)
    image_base64: Optional[str] = None


class VisitOut(BaseModel):
    id: str
    patient_id: str
    symptoms: str
    diagnosis: str
    prescription: str
    raw_text: str = ""
    structured_data: dict = Field(default_factory=dict)
    image_path: Optional[str] = None
    has_image: bool = False
    created_at: datetime

    class Config:
        from_attributes = False


class ScanResult(BaseModel):
    raw_text: str
    structured_data: dict
    confidence: float


class PowerBIData(BaseModel):
    patients: List[dict]
    visits: List[dict]
    summary: dict


class ExportPreviewBody(BaseModel):
    raw_text: str = ""
    structured_data: dict = Field(default_factory=dict)


class BulkPatientsFromScan(BaseModel):
    """Seven cells per inner list: S.NO, PATIENTS NAME, AGE/SEX, MOBILE, DOCTOR, TIMING, REMARKS."""

    rows: List[List[Any]] = Field(default_factory=list)
    raw_text: str = ""
    image_base64: Optional[str] = None
    structured_data: dict = Field(default_factory=dict)


class BulkPatientsFromScanResult(BaseModel):
    created: int
    skipped: int
    patient_ids: List[str]
    visit_id: Optional[str] = None


# ============== Helpers ==============


def patient_to_schema(row: PatientTable) -> Patient:
    return Patient(
        id=str(row.id),
        name=row.name,
        age=row.age,
        gender=row.gender,
        phone=row.phone or "",
        location=row.location or "",
        doctor_name=getattr(row, "doctor_name", None) or "",
        notes=getattr(row, "notes", None) or "",
        created_at=row.created_at,
    )


def visit_to_schema(row: VisitTable) -> VisitOut:
    data = row.structured_data if row.structured_data is not None else {}
    return VisitOut(
        id=str(row.id),
        patient_id=str(row.patient_id),
        symptoms=row.symptoms or "",
        diagnosis=row.diagnosis or "",
        prescription=row.prescription or "",
        raw_text=row.raw_text or "",
        structured_data=dict(data) if isinstance(data, dict) else {},
        image_path=row.image_path,
        has_image=bool(row.image_path),
        created_at=row.created_at,
    )


def decode_base64_image(b64: str) -> bytes:
    if "," in b64:
        b64 = b64.split(",")[1]
    return base64.b64decode(b64)


def save_visit_image(visit_id: uuid.UUID, image_base64: Optional[str]) -> Optional[str]:
    if not image_base64:
        return None
    try:
        UPLOADS_VISITS.mkdir(parents=True, exist_ok=True)
        path = UPLOADS_VISITS / f"{visit_id}.jpg"
        path.write_bytes(decode_base64_image(image_base64))
        rel = f"uploads/visits/{visit_id}.jpg"
        logger.info("Saved visit image %s", rel)
        return rel
    except Exception as e:
        logger.error("Failed to save visit image: %s", e)
        raise HTTPException(status_code=400, detail=f"Invalid image data: {e}")


APPOINTMENT_TABLE_HEADERS: tuple[str, ...] = (
    "S.NO",
    "PATIENTS NAME",
    "AGE/SEX",
    "MOBILE NUMBER",
    "DOCTOR NAME",
    "TIMING",
    "REMARKS",
)

OPENAI_APPOINTMENT_PROMPT = """You digitize a clinic "NEWGEN DAILY APPOINTMENTS" style register: printed column grid + blue/black ink handwriting. Photo may be slightly angled or shadowed.

Columns (left → right), exactly 7 cells per patient row:
1. S.NO — "01", "02", … or "" if that row has no serial (common at bottom of page).
2. PATIENTS NAME — full handwritten name; never leave blank if any name-shaped text appears in that column for this row.
3. AGE/SEX — pattern like 7/M, 43/F, 22/F; "" only if that cell is truly empty.
4. MOBILE NUMBER — Indian 10-digit mobile, digits only in output (no spaces). Trace grid: a long number belongs here, not in NAME.
5. DOCTOR NAME — includes "Dr." and surname; may wrap (e.g. Dr. Suresh Paulraj).
6. TIMING — e.g. 6:30 PM, 07:00 PM, 8:00 PM; "" if blank.
7. REMARKS — short notes (e.g. X-ray, DONE); "" if blank.

Critical rules:
- Follow the printed vertical rules: align each ink stroke with its column. If handwriting drifts, still assign values to the intended column by position.
- One output row = one patient entry. Large blank gaps between groups on the sheet are not separate rows.
- Bottom section often has NAME + AGE/SEX + MOBILE + DOCTOR but missing S.NO and/or TIMING — still output that row with "" for missing fields, never drop NAME or phone if visible.
- Re-read each row before finalizing: if MOBILE is filled but NAME looks empty, zoom mentally on the name cell—faint or small writing is often there.
- Normalize mobile to 10 digits; if you see 11 digits with a leading 0, prefer the standard 10-digit form.
- Do not merge two patients into one row. Do not skip a row because one field is hard to read—best-effort string or "".

Return ONLY valid JSON (no markdown):
{
  "plain_text": "One line per patient for quick reading, or empty string",
  "rows": [["01","Yashmith","7/M","9094696605","Dr. Hemalatha","6:30 PM",""], ...]
}

Each inner array MUST have exactly 7 strings."""


def empty_scan_structured() -> dict[str, Any]:
    return {
        "name": "Unknown",
        "age": 0,
        "gender": "Unknown",
        "phone": "",
        "symptoms": "",
        "diagnosis": "",
        "prescription": "",
        "appointment_table": {
            "headers": list(APPOINTMENT_TABLE_HEADERS),
            "rows": [],
        },
    }


def build_export_document(
    *,
    visit_id: str,
    patient_id: str,
    created_at: datetime,
    raw_text: str,
    structured_data: dict,
    patient_name: Optional[str] = None,
) -> dict[str, Any]:
    return {
        "visit_id": visit_id,
        "patient_id": patient_id,
        "patient_name": patient_name or "",
        "created_at": created_at.isoformat() if created_at else "",
        "raw_text": raw_text,
        "structured_data": structured_data,
    }


def _appointment_export_table(
    payload: dict[str, Any],
) -> Optional[tuple[list[str], list[list[str]]]]:
    sd = payload.get("structured_data")
    if not isinstance(sd, dict):
        return None
    at = sd.get("appointment_table")
    if not isinstance(at, dict):
        return None
    headers = at.get("headers")
    rows = at.get("rows")
    if isinstance(headers, list) and headers:
        hdr_list = [str(h) for h in headers]
    else:
        hdr_list = list(APPOINTMENT_TABLE_HEADERS)
    if not isinstance(rows, list) or not rows:
        return None
    w = len(hdr_list)
    norm: list[list[str]] = []
    for r in rows:
        if not isinstance(r, (list, tuple)):
            continue
        cells = ["" if c is None else str(c) for c in r]
        while len(cells) < w:
            cells.append("")
        norm.append(cells[:w])
    return (hdr_list, norm) if norm else None


def export_to_json_bytes(payload: dict[str, Any]) -> bytes:
    t = _appointment_export_table(payload)
    if t:
        headers, rows = t
        doc = {"headers": headers, "rows": rows}
        return json.dumps(doc, indent=2, ensure_ascii=False).encode("utf-8")
    raw = payload.get("raw_text")
    doc = {"raw_text": "" if raw is None else str(raw)}
    return json.dumps(doc, indent=2, ensure_ascii=False).encode("utf-8")


def export_to_xml_bytes(payload: dict[str, Any]) -> bytes:
    t = _appointment_export_table(payload)
    if t:
        headers, rows = t
        root = ET.Element("appointment_sheet")
        hdr_el = ET.SubElement(root, "headers")
        for h in headers:
            col = ET.SubElement(hdr_el, "column")
            col.text = str(h)
        rows_el = ET.SubElement(root, "rows")
        for row_cells in rows:
            r_el = ET.SubElement(rows_el, "row")
            for i, cell in enumerate(row_cells):
                field = ET.SubElement(r_el, "field")
                if i < len(headers):
                    field.set("column", str(headers[i]))
                field.text = str(cell) if cell is not None else ""
        xml_str = ET.tostring(root, encoding="unicode")
        return f'<?xml version="1.0" encoding="UTF-8"?>\n{xml_str}'.encode("utf-8")
    raw = payload.get("raw_text")
    text = "" if raw is None else str(raw)
    root = ET.Element("ocr_extract")
    rt = ET.SubElement(root, "raw_text")
    rt.text = text
    xml_str = ET.tostring(root, encoding="unicode")
    return f'<?xml version="1.0" encoding="UTF-8"?>\n{xml_str}'.encode("utf-8")


def export_to_xlsx_bytes(payload: dict[str, Any]) -> bytes:
    t = _appointment_export_table(payload)
    if t:
        headers, rows = t
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appointments"
        ws.append(list(headers))
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_cells in rows:
            ws.append(list(row_cells))
        wrap = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrap
        default_widths = [8, 22, 10, 14, 24, 12, 28]
        for i in range(1, len(headers) + 1):
            w = default_widths[i - 1] if i <= len(default_widths) else 18
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    raw = payload.get("raw_text")
    text = "" if raw is None else str(raw)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extract"
    ws.append(
        [
            "No table rows extracted — raw OCR (retake straighter / brighter, or set EMERGENT_LLM_KEY)"
        ]
    )
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.append([""])
    for line in text.splitlines() if text else [""]:
        ws.append([line])
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap
    ws.column_dimensions["A"].width = 96
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def preprocess_image_for_ocr(image: Image.Image) -> Image.Image:
    """Boost contrast / sharpness and upscale small photos—helps Tesseract on desk shots."""
    if image.mode not in ("RGB", "L"):
        image = image.convert("RGB")
    gray = image.convert("L")
    gray = ImageOps.autocontrast(gray, cutoff=1)
    gray = ImageEnhance.Contrast(gray).enhance(1.45)
    gray = ImageEnhance.Sharpness(gray).enhance(1.2)
    w, h = gray.size
    m = max(w, h)
    if m > 0 and m < 1600:
        scale = 1600 / m
        gray = gray.resize(
            (max(1, int(w * scale)), max(1, int(h * scale))),
            Image.Resampling.LANCZOS,
        )
    return gray


async def extract_text_from_image(image_data: bytes) -> str:
    try:
        image = Image.open(io.BytesIO(image_data))
        image = preprocess_image_for_ocr(image)
        psm = os.environ.get("TESSERACT_PSM", "6").strip() or "6"
        cfg = f"--oem 3 --psm {psm}"
        text = pytesseract.image_to_string(image, lang="eng", config=cfg)
        return text.strip()
    except Exception as e:
        logger.error(f"OCR Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")


def _image_bytes_to_jpeg_data_uri(image_data: bytes) -> str:
    """Encode for OpenAI Vision: mild enhance, cap max side, upscale small photos, high JPEG quality."""
    im = Image.open(io.BytesIO(image_data))
    if im.mode != "RGB":
        im = im.convert("RGB")
    im = ImageEnhance.Contrast(im).enhance(1.1)
    im = ImageEnhance.Sharpness(im).enhance(1.12)
    w, h = im.size
    m = max(w, h)
    max_side = max(512, int(os.environ.get("OPENAI_IMAGE_MAX_SIDE", "3072") or "3072"))
    min_up = max(800, int(os.environ.get("OPENAI_IMAGE_MIN_UPSCALE", "2000") or "2000"))
    if m > max_side:
        s = max_side / m
        im = im.resize(
            (max(1, int(w * s)), max(1, int(h * s))),
            Image.Resampling.LANCZOS,
        )
        w, h = im.size
        m = max(w, h)
    elif 0 < m < min_up:
        s = min_up / m
        im = im.resize(
            (max(1, int(w * s)), max(1, int(h * s))),
            Image.Resampling.LANCZOS,
        )
    q = max(75, min(95, int(os.environ.get("OPENAI_JPEG_QUALITY", "92") or "92")))
    buf = io.BytesIO()
    im.save(buf, format="JPEG", quality=q, optimize=True)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/jpeg;base64,{b64}"


def _normalize_mobile_digits(cell: str) -> str:
    """Strip non-digits; keep 10-digit Indian mobiles (drop single leading 0 if 11 digits)."""
    d = re.sub(r"\D", "", cell or "")
    if len(d) == 11 and d.startswith("0"):
        d = d[1:]
    if len(d) == 10 and d.isdigit():
        return d
    if len(d) > 10 and d.isdigit():
        return d[-10:]
    return (cell or "").strip()


def _post_process_appointment_rows(rows: list[list[str]]) -> list[list[str]]:
    out: list[list[str]] = []
    for r in rows:
        cells = list(r)
        if len(cells) > 3:
            cells[3] = _normalize_mobile_digits(cells[3])
        out.append(cells)
    return out


def pad_appointment_row_cells(cells: list[Any]) -> list[str]:
    w = len(APPOINTMENT_TABLE_HEADERS)
    c = ["" if x is None else str(x).strip() for x in cells]
    while len(c) < w:
        c.append("")
    c = c[:w]
    c[3] = _normalize_mobile_digits(c[3])
    return c


def parse_age_sex_cell(value: str) -> tuple[int, str]:
    v = (value or "").strip()
    if not v:
        return 0, "Unknown"
    m = re.match(r"^(\d+)\s*[/\s]+\s*([MmfF])\.?\s*$", v, re.I)
    if m:
        return int(m.group(1)), (
            "Female" if m.group(2).upper() == "F" else "Male"
        )
    dm = re.search(r"(\d+)", v)
    if dm:
        return int(dm.group(1)), "Unknown"
    return 0, "Unknown"


def build_patient_notes_from_row(s_no: str, timing: str, remarks: str) -> str:
    parts: list[str] = []
    if (s_no or "").strip():
        parts.append(f"Sheet S.No: {(s_no or '').strip()}")
    if (timing or "").strip():
        parts.append(f"Appointment time: {(timing or '').strip()}")
    if (remarks or "").strip():
        parts.append(f"Comments: {(remarks or '').strip()}")
    return "\n".join(parts)


def _table_to_plain_text(table: dict[str, Any]) -> str:
    rows = table.get("rows")
    if not isinstance(rows, list) or not rows:
        return ""
    lines: list[str] = []
    for r in rows:
        if not isinstance(r, (list, tuple)):
            continue
        lines.append(" | ".join("" if c is None else str(c) for c in r))
    return "\n".join(lines)


async def extract_appointment_openai_vision(image_data: bytes) -> tuple[str, dict[str, Any]]:
    from openai import AsyncOpenAI

    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise ValueError("OPENAI_API_KEY is not set")

    model = os.environ.get("OPENAI_VISION_MODEL", "gpt-4o").strip() or "gpt-4o"
    uri = _image_bytes_to_jpeg_data_uri(image_data)
    client = AsyncOpenAI(api_key=api_key)
    detail = (os.environ.get("OPENAI_IMAGE_DETAIL", "high") or "high").strip().lower()
    if detail not in ("high", "low", "auto"):
        detail = "high"
    resp = await client.chat.completions.create(
        model=model,
        temperature=0.1,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": OPENAI_APPOINTMENT_PROMPT},
                    {
                        "type": "image_url",
                        "image_url": {"url": uri, "detail": detail},
                    },
                ],
            }
        ],
        response_format={"type": "json_object"},
        max_tokens=8192,
    )
    raw = (resp.choices[0].message.content or "").strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        data = json.loads(_strip_code_fence(raw))
    plain = str(data.get("plain_text") or "").strip()
    rows_in = data.get("rows")
    w = len(APPOINTMENT_TABLE_HEADERS)
    norm: list[list[str]] = []
    if isinstance(rows_in, list):
        for r in rows_in:
            if not isinstance(r, (list, tuple)):
                continue
            cells = ["" if c is None else str(c).strip() for c in r]
            while len(cells) < w:
                cells.append("")
            norm.append(cells[:w])
    norm = _post_process_appointment_rows(norm)
    return plain, {
        "headers": list(APPOINTMENT_TABLE_HEADERS),
        "rows": norm,
    }


def _line_looks_like_table_header(line: str) -> bool:
    u = line.upper()
    hits = sum(
        1
        for k in (
            "S.NO",
            "PATIENT",
            "AGE/SEX",
            "MOBILE",
            "DOCTOR",
            "TIMING",
            "REMARKS",
            "APPOINTMENT",
            "NEWGEN",
        )
        if k in u
    )
    return hits >= 2


def _split_name_agesex(blob: str) -> tuple[str, str]:
    blob = blob.strip()
    if not blob:
        return "", ""
    m = re.search(r"([\d]+\s*/\s*[MFmui]\.?)\s*$", blob, re.I)
    if m:
        agesex = re.sub(r"\s+", "", m.group(1)).replace("M.", "M").replace("F.", "F")
        return blob[: m.start()].strip(), agesex
    return blob, ""


def _split_doctor_timing_remarks(right: str) -> tuple[str, str, str]:
    right = right.strip()
    if not right:
        return "", "", ""
    tm = re.search(
        r"\b(\d{1,2}:\d{2}\s*[APap][Mm]|\d{2}:\d{2}\s*[APap][Mm])\b",
        right,
    )
    if tm:
        timing = re.sub(r"\s+", " ", tm.group(1).strip())
        before = right[: tm.start()].strip()
        after = right[tm.end() :].strip()
        return before, timing, after
    return right, "", ""


def _extract_10digit_phone_span(line: str) -> Optional[tuple[str, int, int]]:
    """Find exactly 10 digits (ignoring spaces/punctuation between them) and its span."""
    positions: list[tuple[int, str]] = [(i, c) for i, c in enumerate(line) if c.isdigit()]
    if len(positions) < 10:
        return None
    digits = "".join(c for _, c in positions)
    for off in range(0, len(digits) - 9):
        chunk = digits[off : off + 10]
        if not chunk.isdigit():
            continue
        start = positions[off][0]
        end = positions[off + 9][0] + 1
        return chunk, start, end
    return None


def _try_parse_pipe_appointment_line(line: str) -> Optional[list[str]]:
    """OCR often reads vertical rules as '|'; join multi-column handwriting into 7 cells."""
    if "|" not in line:
        return None
    parts = [p.strip() for p in line.split("|")]
    while parts and not parts[0]:
        parts.pop(0)
    while parts and not parts[-1]:
        parts.pop()
    if sum(1 for p in parts if p) < 3:
        return None
    if not any(len(re.sub(r"\D", "", p)) >= 10 for p in parts):
        return None
    if len(parts) > 7:
        parts = parts[:6] + [" ".join(parts[6:]).strip()]
    while len(parts) < 7:
        parts.append("")
    return [parts[i] if i < len(parts) else "" for i in range(7)]


def parse_appointment_table_basic(raw_text: str) -> dict[str, Any]:
    headers = list(APPOINTMENT_TABLE_HEADERS)
    rows: list[list[str]] = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line or _line_looks_like_table_header(line):
            continue

        pipe_row = _try_parse_pipe_appointment_line(line)
        if pipe_row:
            rows.append(pipe_row)
            continue

        ph = _extract_10digit_phone_span(line)
        if not ph:
            continue
        mobile, p0, p1 = ph
        left = line[:p0].strip()
        right = line[p1:].strip()
        sm = re.match(r"^(\d{1,2})\s+(.*)$", left)
        if sm:
            sno, rest_left = sm.group(1), sm.group(2)
        else:
            sno, rest_left = "", left
        name, agesex = _split_name_agesex(rest_left)
        doctor, timing, remarks = _split_doctor_timing_remarks(right)
        rows.append([sno, name, agesex, mobile, doctor, timing, remarks])
    return {"headers": headers, "rows": rows}


def _strip_code_fence(text: str) -> str:
    t = text.strip()
    if t.startswith("```json"):
        t = t[7:]
    if t.startswith("```"):
        t = t[3:]
    if t.endswith("```"):
        t = t[:-3]
    return t.strip()


async def extract_appointment_table(raw_text: str) -> dict[str, Any]:
    emergent_key = os.environ.get("EMERGENT_LLM_KEY")
    if emergent_key:
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage

            chat = (
                LlmChat(
                    api_key=emergent_key,
                    session_id=f"clinic-appt-{uuid.uuid4()}",
                    system_message="""You read OCR text from a paper clinic appointment register titled like "DAILY APPOINTMENTS".
Each handwritten row has up to 7 cells in this exact order:
1) S.NO (serial, may be blank for some rows)
2) PATIENTS NAME
3) AGE/SEX (e.g. 7/M., 22/F)
4) MOBILE NUMBER (10 digits)
5) DOCTOR NAME (e.g. Dr. Hemalatha)
6) TIMING (e.g. 6:30 PM, 07:00 PM — may be blank)
7) REMARKS (may be blank)

Respond with ONLY valid JSON of this shape:
{"rows": [["", "name", "age/sex", "phone", "doctor", "time", "remarks"], ...]}

Rules:
- Each inner array MUST have exactly 7 string values (use "" for empty cells).
- Do not include column titles or narrations outside the JSON.
- Copy text as given by OCR; normalize only obvious OCR spacing.
- Include every data row you see; skip blank spacer lines and the printed header row.""",
                )
                .with_model("openai", "gpt-5.2")
            )
            user_message = UserMessage(
                text=f"OCR text from the appointment sheet:\n\n{raw_text}\n\nReturn only the JSON object."
            )
            response = await chat.send_message(user_message)
            data = json.loads(_strip_code_fence(response))
            llm_rows = data.get("rows")
            if not isinstance(llm_rows, list):
                raise ValueError("missing rows array")
            w = len(APPOINTMENT_TABLE_HEADERS)
            norm: list[list[str]] = []
            for r in llm_rows:
                if not isinstance(r, (list, tuple)):
                    continue
                cells = ["" if c is None else str(c).strip() for c in r]
                while len(cells) < w:
                    cells.append("")
                norm.append(cells[:w])
            if norm:
                return {"headers": list(APPOINTMENT_TABLE_HEADERS), "rows": norm}
        except Exception as e:
            logger.warning("Appointment LLM extraction failed, using heuristics: %s", e)
    return parse_appointment_table_basic(raw_text)


def parse_text_basic(raw_text: str) -> dict:
    lines = raw_text.split("\n")
    result = {
        "name": "Unknown",
        "age": 0,
        "gender": "Unknown",
        "phone": "",
        "symptoms": "",
        "diagnosis": "",
        "prescription": "",
    }

    for line in lines:
        line_lower = line.lower()
        if "name" in line_lower:
            result["name"] = line.split(":")[-1].strip() if ":" in line else line.strip()
        elif "age" in line_lower:
            numbers = re.findall(r"\d+", line)
            if numbers:
                result["age"] = int(numbers[0])
        elif "gender" in line_lower or "sex" in line_lower:
            if "male" in line_lower:
                result["gender"] = "Male"
            elif "female" in line_lower:
                result["gender"] = "Female"
        elif "phone" in line_lower or "mobile" in line_lower or "contact" in line_lower:
            numbers = re.findall(r"[\d\-\+\s]{10,}", line)
            if numbers:
                result["phone"] = numbers[0].strip()
        elif "symptom" in line_lower:
            result["symptoms"] = line.split(":")[-1].strip() if ":" in line else ""
        elif "diagnos" in line_lower:
            result["diagnosis"] = line.split(":")[-1].strip() if ":" in line else ""
        elif (
            "prescription" in line_lower
            or "medicine" in line_lower
            or "rx" in line_lower
        ):
            result["prescription"] = (
                line.split(":")[-1].strip() if ":" in line else ""
            )

    return result


async def structure_scan_document(raw_text: str) -> dict:
    table = await extract_appointment_table(raw_text)
    legacy = parse_text_basic(raw_text)
    legacy["appointment_table"] = table
    return legacy


def scan_result_confidence(structured_data: dict) -> float:
    at = structured_data.get("appointment_table")
    if isinstance(at, dict):
        rows = at.get("rows")
        headers = at.get("headers")
        if isinstance(rows, list) and rows:
            w = (
                len(headers)
                if isinstance(headers, list) and headers
                else len(APPOINTMENT_TABLE_HEADERS)
            )
            w = w or len(APPOINTMENT_TABLE_HEADERS)
            filled = 0
            total = 0
            for r in rows:
                if not isinstance(r, (list, tuple)):
                    continue
                cells = ["" if x is None else str(x).strip() for x in r]
                cells.extend([""] * (w - len(cells)))
                for c in cells[:w]:
                    total += 1
                    if c:
                        filled += 1
            if total:
                return round(min(100.0, filled / total * 100), 2)
    keys = [k for k in structured_data.keys() if k != "appointment_table"]
    filled_fields = sum(
        1
        for k in keys
        if structured_data.get(k) not in (None, "", "Unknown", 0)
    )
    return round(min(100.0, filled_fields / max(len(keys), 1) * 100), 2)


async def scan_from_image_data(image_data: bytes) -> ScanResult:
    """Use OpenAI Vision for the appointment table when OPENAI_API_KEY is set; else Tesseract."""
    if os.environ.get("OPENAI_API_KEY", "").strip():
        try:
            plain, table = await extract_appointment_openai_vision(image_data)
            rows = table.get("rows") if isinstance(table.get("rows"), list) else []
            if rows:
                legacy = parse_text_basic(plain)
                legacy["appointment_table"] = table
                raw_text = plain if plain else _table_to_plain_text(table)
                confidence = scan_result_confidence(legacy)
                return ScanResult(
                    raw_text=raw_text,
                    structured_data=legacy,
                    confidence=confidence,
                )
            logger.info("OpenAI vision returned no rows; falling back to Tesseract")
        except Exception as e:
            logger.warning("OpenAI vision table extraction failed: %s", e)

    raw_text = await extract_text_from_image(image_data)
    if not raw_text.strip():
        return ScanResult(
            raw_text="",
            structured_data=empty_scan_structured(),
            confidence=0.0,
        )
    structured_data = await structure_scan_document(raw_text)
    return ScanResult(
        raw_text=raw_text,
        structured_data=structured_data,
        confidence=scan_result_confidence(structured_data),
    )


# ============== API: root, health, scan ==============


@api_router.get("/")
async def root():
    return {"message": "Clinic Digitization API", "version": "1.0"}


@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


@api_router.post("/scan", response_model=ScanResult)
async def scan_image(image: UploadFile = File(...)):
    logger.info("Received scan request for file: %s", image.filename)
    image_data = await image.read()
    return await scan_from_image_data(image_data)


@api_router.post("/scan/base64", response_model=ScanResult)
async def scan_image_base64(data: dict):
    base64_string = data.get("image_base64", "")
    if not base64_string:
        raise HTTPException(status_code=400, detail="No image data provided")
    if "," in base64_string:
        base64_string = base64_string.split(",")[1]
    try:
        image_data = base64.b64decode(base64_string)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid base64 data: {str(e)}")
    return await scan_from_image_data(image_data)


# ============== Patients ==============


@api_router.post("/patients", response_model=Patient)
async def create_patient(patient: PatientCreate, db: AsyncSession = Depends(get_db)):
    row = PatientTable(
        name=patient.name,
        age=patient.age,
        gender=patient.gender,
        phone=patient.phone or "",
        doctor_name=patient.doctor_name or "",
        notes=patient.notes or "",
        location=patient.location or None,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return patient_to_schema(row)


@api_router.post("/patients/bulk-from-scan", response_model=BulkPatientsFromScanResult)
async def bulk_patients_from_scan(
    body: BulkPatientsFromScan, db: AsyncSession = Depends(get_db)
):
    patient_ids: list[str] = []
    skipped = 0
    for _, raw_cells in enumerate(body.rows):
        if not isinstance(raw_cells, (list, tuple)):
            skipped += 1
            continue
        cells = pad_appointment_row_cells(list(raw_cells))
        name = cells[1].strip()
        phone = cells[3].strip()
        if not name and not phone:
            skipped += 1
            continue
        display_name = (
            name
            if name
            else (
                f"Patient {phone[-4:]}"
                if len(phone) >= 4
                else "Unknown patient"
            )
        )
        age, gender = parse_age_sex_cell(cells[2])
        doctor = cells[4].strip()
        timing = cells[5]
        remarks = cells[6]
        notes = build_patient_notes_from_row(cells[0], timing, remarks)
        prow = PatientTable(
            name=display_name,
            age=age,
            gender=gender,
            phone=phone or "",
            doctor_name=doctor,
            notes=notes,
            location=None,
        )
        db.add(prow)
        await db.flush()
        patient_ids.append(str(prow.id))

    visit_id_out: Optional[str] = None
    if patient_ids:
        base_sd: dict[str, Any] = {}
        if isinstance(body.structured_data, dict):
            base_sd = dict(body.structured_data)
        at = base_sd.get("appointment_table")
        if not isinstance(at, dict) or not isinstance(at.get("rows"), list):
            base_sd["appointment_table"] = {
                "headers": list(APPOINTMENT_TABLE_HEADERS),
                "rows": [
                    pad_appointment_row_cells(list(r))
                    for r in body.rows
                    if isinstance(r, (list, tuple))
                ],
            }
        base_sd["visit_kind"] = "appointment_sheet_batch"
        base_sd["patient_ids_created"] = patient_ids

        pid_first = uuid.UUID(patient_ids[0])
        vid = uuid.uuid4()
        img_path = save_visit_image(vid, body.image_base64)
        vrow = VisitTable(
            id=vid,
            patient_id=pid_first,
            symptoms="",
            diagnosis="Appointment sheet scan",
            prescription=(
                f"{len(patient_ids)} patient profile(s) created from this sheet."
            ),
            raw_text=body.raw_text or "",
            structured_data=base_sd,
            image_path=img_path,
        )
        db.add(vrow)
        visit_id_out = str(vid)

    await db.commit()
    return BulkPatientsFromScanResult(
        created=len(patient_ids),
        skipped=skipped,
        patient_ids=patient_ids,
        visit_id=visit_id_out,
    )


@api_router.get("/patients", response_model=List[Patient])
async def list_patients(
    db: AsyncSession = Depends(get_db),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    search: Optional[str] = None,
):
    q = select(PatientTable).offset(skip).limit(limit).order_by(PatientTable.created_at.desc())
    if search:
        term = f"%{search}%"
        q = q.where(
            or_(
                PatientTable.name.ilike(term),
                PatientTable.phone.ilike(term),
                PatientTable.location.ilike(term),
                PatientTable.doctor_name.ilike(term),
                PatientTable.notes.ilike(term),
            )
        )
    r = await db.execute(q)
    rows = r.scalars().all()
    return [patient_to_schema(p) for p in rows]


@api_router.get("/patients/{patient_id}", response_model=Patient)
async def get_patient(patient_id: str, db: AsyncSession = Depends(get_db)):
    try:
        pid = uuid.UUID(patient_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Patient not found")
    row = await db.get(PatientTable, pid)
    if not row:
        raise HTTPException(status_code=404, detail="Patient not found")
    return patient_to_schema(row)


@api_router.put("/patients/{patient_id}", response_model=Patient)
async def update_patient(
    patient_id: str, patient_update: PatientCreate, db: AsyncSession = Depends(get_db)
):
    try:
        pid = uuid.UUID(patient_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Patient not found")
    row = await db.get(PatientTable, pid)
    if not row:
        raise HTTPException(status_code=404, detail="Patient not found")
    row.name = patient_update.name
    row.age = patient_update.age
    row.gender = patient_update.gender
    row.phone = patient_update.phone or ""
    row.doctor_name = patient_update.doctor_name or ""
    row.notes = patient_update.notes or ""
    row.location = patient_update.location or None
    await db.commit()
    await db.refresh(row)
    return patient_to_schema(row)


@api_router.delete("/patients/{patient_id}")
async def delete_patient(patient_id: str, db: AsyncSession = Depends(get_db)):
    try:
        pid = uuid.UUID(patient_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Patient not found")
    row = await db.get(PatientTable, pid)
    if not row:
        raise HTTPException(status_code=404, detail="Patient not found")

    visit_result = await db.execute(select(VisitTable).where(VisitTable.patient_id == pid))
    for v in visit_result.scalars().all():
        if v.image_path:
            fp = ROOT_DIR / v.image_path
            if fp.is_file():
                try:
                    fp.unlink()
                except OSError:
                    pass

    await db.execute(delete(VisitTable).where(VisitTable.patient_id == pid))
    await db.delete(row)
    await db.commit()
    return {"message": "Patient deleted successfully"}


# ============== Visits ==============


@api_router.post("/visits", response_model=VisitOut)
async def create_visit(visit: VisitCreate, db: AsyncSession = Depends(get_db)):
    try:
        pid = uuid.UUID(visit.patient_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Patient not found")
    patient_row = await db.get(PatientTable, pid)
    if not patient_row:
        raise HTTPException(status_code=404, detail="Patient not found")

    vid = uuid.uuid4()
    image_path = save_visit_image(vid, visit.image_base64)

    row = VisitTable(
        id=vid,
        patient_id=pid,
        symptoms=visit.symptoms or "",
        diagnosis=visit.diagnosis or "",
        prescription=visit.prescription or "",
        raw_text=visit.raw_text or "",
        structured_data=visit.structured_data or {},
        image_path=image_path,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    logger.info("Created visit %s for patient %s", row.id, visit.patient_id)
    return visit_to_schema(row)


@api_router.get("/visits", response_model=List[VisitOut])
async def list_visits(
    db: AsyncSession = Depends(get_db),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    patient_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    q = select(VisitTable).order_by(VisitTable.created_at.desc()).offset(skip).limit(limit)
    if patient_id:
        try:
            q = q.where(VisitTable.patient_id == uuid.UUID(patient_id))
        except ValueError:
            pass
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
            q = q.where(VisitTable.created_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
            q = q.where(VisitTable.created_at <= end_dt)
        except ValueError:
            pass
    r = await db.execute(q)
    rows = r.scalars().all()
    return [visit_to_schema(v) for v in rows]


@api_router.get("/visits/{visit_id}", response_model=VisitOut)
async def get_visit(visit_id: str, db: AsyncSession = Depends(get_db)):
    try:
        vid = uuid.UUID(visit_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Visit not found")
    row = await db.get(VisitTable, vid)
    if not row:
        raise HTTPException(status_code=404, detail="Visit not found")
    return visit_to_schema(row)


@api_router.put("/visits/{visit_id}", response_model=VisitOut)
async def update_visit(
    visit_id: str, visit_update: VisitCreate, db: AsyncSession = Depends(get_db)
):
    try:
        vid = uuid.UUID(visit_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Visit not found")
    row = await db.get(VisitTable, vid)
    if not row:
        raise HTTPException(status_code=404, detail="Visit not found")
    try:
        pid = uuid.UUID(visit_update.patient_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Patient not found")
    patient_row = await db.get(PatientTable, pid)
    if not patient_row:
        raise HTTPException(status_code=404, detail="Patient not found")

    row.patient_id = pid
    row.symptoms = visit_update.symptoms or ""
    row.diagnosis = visit_update.diagnosis or ""
    row.prescription = visit_update.prescription or ""
    row.raw_text = visit_update.raw_text or ""
    row.structured_data = visit_update.structured_data or {}
    if visit_update.image_base64:
        row.image_path = save_visit_image(row.id, visit_update.image_base64)
    await db.commit()
    await db.refresh(row)
    return visit_to_schema(row)


@api_router.delete("/visits/{visit_id}")
async def delete_visit(visit_id: str, db: AsyncSession = Depends(get_db)):
    try:
        vid = uuid.UUID(visit_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Visit not found")
    row = await db.get(VisitTable, vid)
    if not row:
        raise HTTPException(status_code=404, detail="Visit not found")
    if row.image_path:
        fp = ROOT_DIR / row.image_path
        if fp.is_file():
            fp.unlink()
    await db.delete(row)
    await db.commit()
    return {"message": "Visit deleted successfully"}


@api_router.get("/visits/{visit_id}/image")
async def get_visit_image(visit_id: str, db: AsyncSession = Depends(get_db)):
    try:
        vid = uuid.UUID(visit_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Image not found")
    row = await db.get(VisitTable, vid)
    if not row or not row.image_path:
        raise HTTPException(status_code=404, detail="Image not found")
    path = ROOT_DIR / row.image_path
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Image not found")
    return FileResponse(path, media_type="image/jpeg")


def _export_response(
    content: bytes, media_type: str, filename: str
) -> Response:
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/visits/{visit_id}/export")
async def export_visit(
    visit_id: str,
    db: AsyncSession = Depends(get_db),
    format: str = Query("json", description="json, xml, or xlsx"),
):
    try:
        vid = uuid.UUID(visit_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Visit not found")
    row = await db.get(VisitTable, vid)
    if not row:
        raise HTTPException(status_code=404, detail="Visit not found")
    patient = await db.get(PatientTable, row.patient_id)
    patient_name = patient.name if patient else ""
    payload = build_export_document(
        visit_id=str(row.id),
        patient_id=str(row.patient_id),
        created_at=row.created_at,
        raw_text=row.raw_text or "",
        structured_data=row.structured_data or {}
        if isinstance(row.structured_data, dict)
        else {},
        patient_name=patient_name,
    )
    fmt = format.lower().strip()
    if fmt == "json":
        body = export_to_json_bytes(payload)
        return _export_response(body, "application/json", f"visit-{visit_id}.json")
    if fmt in ("xml",):
        body = export_to_xml_bytes(payload)
        return _export_response(body, "application/xml", f"visit-{visit_id}.xml")
    if fmt in ("xlsx", "xls", "excel"):
        body = export_to_xlsx_bytes(payload)
        return _export_response(
            body,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"visit-{visit_id}.xlsx",
        )
    raise HTTPException(status_code=400, detail="format must be json, xml, or xlsx")


@api_router.post("/export/preview")
async def export_preview(
    body: ExportPreviewBody,
    format: str = Query("json", description="json, xml, or xlsx"),
):
    payload = build_export_document(
        visit_id="preview",
        patient_id="",
        created_at=datetime.utcnow(),
        raw_text=body.raw_text or "",
        structured_data=body.structured_data or {},
        patient_name="",
    )
    fmt = format.lower().strip()
    if fmt == "json":
        b = export_to_json_bytes(payload)
        return _export_response(b, "application/json", "visit-preview.json")
    if fmt in ("xml",):
        b = export_to_xml_bytes(payload)
        return _export_response(b, "application/xml", "visit-preview.xml")
    if fmt in ("xlsx", "xls", "excel"):
        b = export_to_xlsx_bytes(payload)
        return _export_response(
            b,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "visit-preview.xlsx",
        )
    raise HTTPException(status_code=400, detail="format must be json, xml, or xlsx")


@api_router.get("/powerbi/data", response_model=PowerBIData)
async def get_powerbi_data(db: AsyncSession = Depends(get_db)):
    pr = await db.execute(select(PatientTable))
    patients = pr.scalars().all()
    vr = await db.execute(select(VisitTable))
    visits = vr.scalars().all()

    total_patients = len(patients)
    total_visits = len(visits)

    age_groups = {"0-18": 0, "19-40": 0, "41-60": 0, "60+": 0}
    for p in patients:
        age = p.age or 0
        if age <= 18:
            age_groups["0-18"] += 1
        elif age <= 40:
            age_groups["19-40"] += 1
        elif age <= 60:
            age_groups["41-60"] += 1
        else:
            age_groups["60+"] += 1

    gender_dist = {"Male": 0, "Female": 0, "Other": 0}
    for p in patients:
        gender = p.gender or "Other"
        if gender in gender_dist:
            gender_dist[gender] += 1
        else:
            gender_dist["Other"] += 1

    symptom_counts: dict[str, int] = {}
    for v in visits:
        symptoms = v.symptoms or ""
        for symptom in symptoms.split(","):
            s = symptom.strip().lower()
            if s:
                symptom_counts[s] = symptom_counts.get(s, 0) + 1

    formatted_patients = []
    for p in patients:
        formatted_patients.append(
            {
                "id": str(p.id),
                "name": p.name,
                "age": p.age,
                "gender": p.gender,
                "phone": p.phone,
                "location": p.location or "",
                "doctor_name": getattr(p, "doctor_name", "") or "",
                "notes": getattr(p, "notes", "") or "",
                "created_at": p.created_at.isoformat()
                if p.created_at
                else "",
            }
        )

    formatted_visits = []
    for v in visits:
        formatted_visits.append(
            {
                "id": str(v.id),
                "patient_id": str(v.patient_id),
                "symptoms": v.symptoms,
                "diagnosis": v.diagnosis,
                "prescription": v.prescription,
                "created_at": v.created_at.isoformat() if v.created_at else "",
            }
        )

    summary = {
        "total_patients": total_patients,
        "total_visits": total_visits,
        "age_distribution": age_groups,
        "gender_distribution": gender_dist,
        "top_symptoms": dict(
            sorted(symptom_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        ),
        "generated_at": datetime.utcnow().isoformat(),
    }

    return PowerBIData(
        patients=formatted_patients, visits=formatted_visits, summary=summary
    )


@api_router.get("/stats")
async def get_dashboard_stats(db: AsyncSession = Depends(get_db)):
    total_patients = (
        await db.execute(select(func.count()).select_from(PatientTable))
    ).scalar_one()
    total_visits = (
        await db.execute(select(func.count()).select_from(VisitTable))
    ).scalar_one()

    today_start = datetime.now(timezone.utc).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    today_visits = (
        await db.execute(
            select(func.count())
            .select_from(VisitTable)
            .where(VisitTable.created_at >= today_start)
        )
    ).scalar_one()

    recent_q = (
        select(VisitTable)
        .order_by(VisitTable.created_at.desc())
        .limit(5)
    )
    recent_r = await db.execute(recent_q)
    recent_rows = recent_r.scalars().all()

    return {
        "total_patients": total_patients,
        "total_visits": total_visits,
        "today_visits": today_visits,
        "recent_visits": [visit_to_schema(v).model_dump(mode="json") for v in recent_rows],
    }


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
